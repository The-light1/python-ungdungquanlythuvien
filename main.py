from msilib.schema import Font
from tkinter import Tk
from tkinter.filedialog import asksaveasfilename

import pyodbc as odbc
from datetime import datetime

import sys
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtWidgets import QApplication, QTableWidgetItem, QMessageBox, QLineEdit, QMainWindow
from PyQt5.QtGui import QKeySequence
from openpyxl.styles import Alignment, Side, Border
from openpyxl.workbook import Workbook
from DangNhap import Ui_DangNhap
from Trangchu import Ui_Trangchu
from muonsach import Ui_muonsach
from thongke import Ui_Thongke
from quanly import Ui_quanly
import openpyxl
#from docxtpl import Dox
conn = odbc.connect("Server=DESKTOP-5QTVEPE\SQLEXPRESS;DATABASE=Baocao;Trusted_Conection=True;PORT=1433;DRIVER={SQL Server}")
count = 1
results = ()
#   login_loading

class Dangnhap:
    def __init__(self):
        self.main_win = QMainWindow()
        self.uic = Ui_DangNhap()
        self.uic.setupUi(self.main_win)
        self.uic.p_dn.clicked.connect(self.check_login)
        self.uic.p_dn.setShortcut(QKeySequence(Qt.Key.Key_Return))
        password_line_edit = self.main_win.findChild(QLineEdit, "txt_mk")
        if password_line_edit:
            password_line_edit.setEchoMode(QLineEdit.Password)

    def show(self):
        self.main_win.show()

    def check_login(self):
        global count
        self.tk = self.uic.txt_tk.text()
        self.mk = self.uic.txt_mk.text()
        if len(self.tk) == 0 or len(self.mk) == 0:
            QMessageBox.critical(self.main_win, "Lỗi!", "Vui lòng điền đủ thông tin đăng nhập")
        else:
            # Kiểm tra xem tài khoản tồn tại
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM Account WHERE Username = ?", self.tk)
            rows = cursor.fetchall()  # Sử dụng fetchall để lấy tất cả dòng dữ liệu
            # Trong phần kiểm tra tài khoản tồn tại
            if rows:
                if count > 3:
                    UPDATE_account = '''UPDATE Account SET TrangThai = ? WHERE Username = ?'''
                    cursor.execute(UPDATE_account, 0, self.tk)
                    conn.commit()
                    QMessageBox.critical(self.main_win, "Cảnh báo!", "Bạn đã nhập sai mật khẩu quá {} lần\n Tài Khoản bị khóa ".format(count))
                    count = 0
                # Tài khoản tồn tại, kiểm tra mật khẩu
                for row in rows:
                    # Truy cập cột "Password" từ tuple
                    if row.Password == self.mk:
                        if row.TrangThai == 1:
                           self.trangchu = Trangchu()
                           self.trangchu.show1()
                           self.main_win.close()
                           break  # Khi đã tìm thấy mật khẩu khớp, thoát vòng lặp
                        else:
                            QMessageBox.critical(self.main_win, "Lỗi!", "Tài khoản cúa bạn đã bị khóa")
                    else:
                        QMessageBox.critical(self.main_win, "Lỗi!", "Sai mật khẩu!")
                        count += 1
            else:
                QMessageBox.critical(self.main_win, "Lỗi!", "Tài khoản không tồn tại!")

class Trangchu:
    def __init__(self):
        self.main_win1 = QMainWindow()
        self.uic = Ui_Trangchu()
        self.uic.setupUi(self.main_win1)
        self.uic.p_muonvatra.clicked.connect(self.trangmuonvatra)
        self.uic.p_quanly.clicked.connect(self.trangql)
        self.uic.p_thongke.clicked.connect(self.trangthongke)
        self.uic.p_dx.clicked.connect(self.trangdn)

    def trangdn(self):
        self.main_win = Dangnhap()
        self.main_win.show()
        self.main_win1.close()

    def trangmuonvatra(self):
        self.main_muon = muonsach()
        self.main_muon.show_muon()
        self.main_win1.close()

    def trangql(self):
        self.main_ql = Quanly()
        self.main_ql.show_ql()
        self.main_win1.close()

    def trangthongke(self):
        self.main_tk = Thongke()
        self.main_tk.show_tk()
        self.main_win1.close()

    def show1(self):
        self.main_win1.show()
class muonsach:
    def __init__(self):
        self.main_win_muon = QMainWindow()
        self.uic = Ui_muonsach()
        self.uic.setupUi(self.main_win_muon)
        self.uic.p_them.clicked.connect(self.them)
        self.uic.p_sua.clicked.connect(self.sua)
        self.uic.p_xoa.clicked.connect(self.xoa)
        self.uic.p_tim.clicked.connect(self.timsach)
        self.uic.table_sach.cellClicked.connect(self.clicked_table)
        self.uic.p_quanly.clicked.connect(self.trangql)
        self.uic.p_trangchu.clicked.connect(self.trangchu)
        self.uic.p_dx.clicked.connect(self.trangdn)
        self.uic.p_thongke.clicked.connect(self.trangtk)

    def trangtk(self):
        self.main_tk = Thongke()
        self.main_tk.show_tk()
        self.main_win_muon.close()

    def trangdn(self):
        self.main_win = Dangnhap()
        self.main_win.show()
        self.main_win_muon.close()

    def trangchu(self):
        self.trangchu = Trangchu()
        self.trangchu.show1()
        self.main_win_muon.close()

    def trangql(self):
        self.main_quanli = Quanly()
        self.main_quanli.show_ql()
        self.main_win_muon.close()

    def load_table_sach(self):
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Muonvatra")
        result = cursor.fetchall()
        self.uic.table_sach.clearContents()  # clear trước khi cập nhật
        rowcount = len(result)
        self.uic.table_sach.setRowCount(rowcount)
        table_row = 0
        for row in result:
            self.uic.table_sach.setItem(table_row, 0, QTableWidgetItem(str(row[0])))
            self.uic.table_sach.setItem(table_row, 1, QTableWidgetItem(str(row[1])))
            self.uic.table_sach.setItem(table_row, 2, QTableWidgetItem(str(row[2])))
            self.uic.table_sach.setItem(table_row, 3, QTableWidgetItem(str(row[3])))
            self.uic.table_sach.setItem(table_row, 4, QTableWidgetItem(str(row[4])))
            self.uic.table_sach.setItem(table_row, 5, QTableWidgetItem(str(row[5])))
            self.uic.table_sach.setItem(table_row, 6, QTableWidgetItem(str(row[6])))
            table_row += 1

    def show_muon(self):
        self.main_win_muon.show()
        self.load_table_sach()

    def them(self):
     try:
        insert_sach = "INSERT INTO Muonvatra ( Tennguoimuon, Masach, Tensach, Soluong, NgayMuon, NgayTra) VALUES ( ?, ?, ?, ?, ?, ?)"
        cursor = conn.cursor()
        ten_nguoi_muon = self.uic.txt_tennguoi.text()
        Ma_sach = self.uic.txt_ms.text()
        Ten_sach = self.uic.txt_tensach.text()
        so_luong = self.uic.txt_soluong.text()
        ngay_muon = self.uic.Ngaymuon.date().toPyDate()
        ngay_muon_sql_format = ngay_muon.strftime('%Y-%m-%d')
        ngay_tra = self.uic.Ngaytra.date().toPyDate()
        ngay_tra_sql_format = ngay_tra.strftime('%Y-%m-%d')
        #lien_he = self.uic.Lienhe.text()
        if not ten_nguoi_muon or not Ma_sach or not Ten_sach or not so_luong or not ngay_muon or not ngay_tra:
            QMessageBox.critical(self.main_win_muon, "Lỗi!", "Vui lòng nhập đầy đủ thông tin sách.")
        elif not Ma_sach.isdigit() or int(Ma_sach) <= 0:
            QMessageBox.critical(self.main_win_muon, "Lỗi!", "Mã sách không hợp lệ, vui lòng nhập lại (chỉ được chứa ký tự số dương).")
        elif not so_luong.isdigit() or int(so_luong) <= 0:
            QMessageBox.critical(self.main_win_muon, "Lỗi!", "Số lượng sách không hợp lệ, vui lòng nhập lại (chỉ được chứa ký tự số dương).")
        else:
            cursor.execute(insert_sach, (ten_nguoi_muon, int(Ma_sach), Ten_sach, int(so_luong), ngay_muon_sql_format, ngay_tra_sql_format))
            conn.commit()
            self.load_table_sach()
            QMessageBox.information(self.main_win_muon, "Thành công", "Thêm sách thành công!")
     except Exception as e:
        QMessageBox.critical(self.main_win_muon, "Lỗi!", f"Lỗi thực thi: {str(e)}")

    def clicked_table(self, row):
        global selected_ID
        selected_ID = self.uic.table_sach.item(row, 0).text()
        ma_ID = self.uic.table_sach.item(row, 0).text() if self.uic.table_sach.item(row, 0) else ""
        ten_nguoi_muon = self.uic.table_sach.item(row, 1).text() if self.uic.table_sach.item(row, 1) else ""
        ma_sach = self.uic.table_sach.item(row, 2).text() if self.uic.table_sach.item(row, 2) else ""
        ten_sach = self.uic.table_sach.item(row, 3).text() if self.uic.table_sach.item(row, 3) else ""
        so_luong = self.uic.table_sach.item(row, 4).text() if self.uic.table_sach.item(row, 4) else ""
        Ngay_muon = self.uic.table_sach.item(row, 5).text() if self.uic.table_sach.item(row, 5) else ""
        Ngay_tra = self.uic.table_sach.item(row, 6).text() if self.uic.table_sach.item(row, 6) else ""

        # Đặt giá trị vào các widget
        self.uic.txt_ID.setText(ma_ID)
        self.uic.txt_tennguoi.setText(ten_nguoi_muon)
        self.uic.txt_ms.setText(ma_sach)
        self.uic.txt_tensach.setText(ten_sach)
        self.uic.txt_soluong.setText(so_luong)
        self.uic.Ngaymuon.setDate(QDate.fromString(Ngay_muon, "yyyy-MM-dd"))
        self.uic.Ngaytra.setDate(QDate.fromString(Ngay_tra, "yyyy-MM-dd"))

    def sua(self):
     try:
        update_muon = '''UPDATE Muonvatra SET Tennguoimuon = ?, Masach = ?, Tensach = ?, Soluong = ?, NgayMuon = ?, NgayTra = ? WHERE ID = ?'''
        cursor = conn.cursor()
        ma_ID = self.uic.txt_ID.text()
        ten_nguoi_muon = self.uic.txt_tennguoi.text()
        ma_sach = self.uic.txt_ms.text()
        ten_sach = self.uic.txt_tensach.text()
        so_luong = self.uic.txt_soluong.text()
        Ngay_muon = self.uic.Ngaymuon.text()
        Ngay_tra = self.uic.Ngaytra.text()
        if ma_ID != selected_ID:
            QMessageBox.critical(self.main_win_muon, "Lỗi!", "Mã ID không thể chỉnh sửa")
            self.uic.txt_ID.setText("{}".format(selected_ID))
        else:
            if int(self.uic.txt_ms.text()) < 0 or int(self.uic.txt_soluong.text()) < 0:
                QMessageBox.critical(self.main_win_muon, "Lỗi!", "Mã sách và số lượng phải là số không âm")
            else:
                ma_sach = self.uic.txt_ms.text()
                so_luong = self.uic.txt_soluong.text()
                cursor.execute(update_muon, ten_nguoi_muon, ma_sach,ten_sach, so_luong, Ngay_muon, Ngay_tra, ma_ID)
                conn.commit()
                self.load_table_sach()
     except Exception as e:
         # Xử lý lỗi nếu có
         QMessageBox.critical(self.main_win_muon, "Lỗi!", f"Lỗi thực thi: {str(e)}")
    def xoa(self):
        delete_Id = '''DELETE FROM Muonvatra WHERE ID = ?'''
        cursor = conn.cursor()
        ma_id = self.uic.txt_ID.text()
        if ma_id:
            cursor.execute(delete_Id, (ma_id,))
            conn.commit()
            self.load_table_sach()
        else:
            QMessageBox.critical(self.main_win_muon, "Lỗi", "Vui lòng nhập số dòng muốn xóa!!")

    def timsach(self):
        timkiem = self.uic.txt_timkiem.text().strip()
        if len(timkiem) == 0:
            self.load_table_sach()
        else:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM Muonvatra Where Tennguoimuon = ?", timkiem)
            result = cursor.fetchall()
            self.uic.table_sach.clearContents()  # clear trước khi cập nhật
            rowcount = len(result)
            self.uic.table_sach.setRowCount(rowcount)
            table_row = 0
            for row in result:
                self.uic.table_sach.setItem(table_row, 0, QTableWidgetItem(str(row[0])))
                self.uic.table_sach.setItem(table_row, 1, QTableWidgetItem(str(row[1])))
                self.uic.table_sach.setItem(table_row, 2, QTableWidgetItem(str(row[2])))
                self.uic.table_sach.setItem(table_row, 3, QTableWidgetItem(str(row[3])))
                self.uic.table_sach.setItem(table_row, 4, QTableWidgetItem(str(row[4])))
                self.uic.table_sach.setItem(table_row, 5, QTableWidgetItem(str(row[5])))
                self.uic.table_sach.setItem(table_row, 6, QTableWidgetItem(str(row[6])))
                table_row += 1
        self.uic.txt_timkiem.setText("")
class Quanly:
    def __init__(self):
        self.main_win_quanli= QMainWindow()
        self.uic = Ui_quanly()
        self.uic.setupUi(self.main_win_quanli)
        self.uic.p_them.clicked.connect(self.themsach)
        self.uic.p_sua.clicked.connect(self.suasach)
        self.uic.p_xoa.clicked.connect(self.xoasach)
        self.uic.p_tim.clicked.connect(self.timsach)
        self.uic.tb_quanly.cellClicked.connect(self.clicked_table_ql)
        self.uic.p_thongke.clicked.connect(self.trangtk)
        self.uic.p_trangchu.clicked.connect(self.trangchu)
        self.uic.p_muonvatra.clicked.connect(self.trangmuon)
        self.uic.p_dx.clicked.connect(self.trangdn)

    def trangdn(self):
        self.main_win = Dangnhap()
        self.main_win.show()
        self.main_win_quanli.close()

    def trangmuon(self):
        self.main_win_muon = muonsach()
        self.main_win_muon.show_muon()
        self.main_win_quanli.close()

    def trangchu(self):
        self.maintrangchu = Trangchu()
        self.maintrangchu.show1()
        self.main_win_quanli.close()

    def trangtk(self):
        self.main_thongke = Thongke ()
        self.main_thongke.show_tk()
        self.main_win_quanli.close()

    def show_ql(self):
        self.main_win_quanli.show()
        self.load_table_ql()

    def clicked_table_ql(self, row):
        # Lấy giá trị
        global selected_ms
        selected_ms = self.uic.tb_quanly.item(row, 0).text()
        ma_sach = self.uic.tb_quanly.item(row, 0).text()
        ten_sach = self.uic.tb_quanly.item(row, 1).text()
        soluong = self.uic.tb_quanly.item(row, 2).text()
        nha_xuat_ban = self.uic.tb_quanly.item(row, 3).text()
        Tac_gia = self.uic.tb_quanly.item(row, 4).text()

        # Đặt giá trị
        self.uic.txt_ms.setText(ma_sach)
        self.uic.txt_tensach.setText(ten_sach)
        self.uic.txt_soluong.setText(soluong)
        self.uic.txt_Nhaxuatban.setText(nha_xuat_ban)
        self.uic.txt_tacgia.setText(Tac_gia)

    def load_table_ql(self):
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Quanly")
        result = cursor.fetchall()
        self.uic.tb_quanly.clearContents()  # clear trước khi cập nhật
        rowcount = len(result)
        self.uic.tb_quanly.setRowCount(rowcount)
        table_row = 0
        for row in result:
            self.uic.tb_quanly.setItem(table_row, 0, QTableWidgetItem(str(row[0])))
            self.uic.tb_quanly.setItem(table_row, 1, QTableWidgetItem(str(row[1])))
            self.uic.tb_quanly.setItem(table_row, 2, QTableWidgetItem(str(row[2])))
            self.uic.tb_quanly.setItem(table_row, 3, QTableWidgetItem(str(row[3])))
            self.uic.tb_quanly.setItem(table_row, 4, QTableWidgetItem(str(row[4])))

            table_row += 1

    def themsach(self):
     try:
        insert_nv = "INSERT INTO Quanly ( Tensach, Soluong, Nxb, Tacgia) VALUES ( ?, ?, ?, ?)"
        cursor = conn.cursor()
        ten_sach = self.uic.txt_tensach.text()
        So_luong = self.uic.txt_soluong.text()
        Nxb = self.uic.txt_Nhaxuatban.text()
        Tac_gia = self.uic.txt_tacgia.text()
        if not ten_sach or not So_luong or not Nxb or not Tac_gia:
            QMessageBox.critical(self.main_win_quanli, "Lỗi!", "Vui lòng nhập thông tin sách để thêm.")
            self.load_table_ql()
        else:
            cursor.execute(insert_nv,  ten_sach, So_luong, Nxb, Tac_gia)
            conn.commit()
            self.load_table_ql()
     except Exception as e:
     # Xử lý lỗi nếu có
        QMessageBox.critical(self.main_win_quanli, "Lỗi!", f"Lỗi thực thi: {str(e)}")
    def suasach(self):
        update_sach= '''UPDATE Quanly SET Tensach = ?, Soluong = ?, Nxb = ?, Tacgia = ? WHERE  Masach = ?'''
        cursor = conn.cursor()
        ma_sach = self.uic.txt_ms.text().strip()
        ten_sach = self.uic.txt_tensach.text().strip()
        so_luong = self.uic.txt_soluong.text().strip()
        nxb = self.uic.txt_Nhaxuatban.text().strip()
        tac_gia = self.uic.txt_tacgia.text().strip()

        # Kiểm tra xem dữ liệu có hợp lệ không
        if not ma_sach or not ten_sach or not so_luong or not nxb or not tac_gia:
            QMessageBox.critical(self.main_win_quanli, "Lỗi!", "Vui lòng nhập đầy đủ thông tin để sửa sách.")
            return
        if not so_luong.isdigit() or int(so_luong) <= 0:
            QMessageBox.critical(self.main_win_quanli, "Lỗi!", "Số lượng phải là số dương.")
            return

        # Kiểm tra nếu mã sách không khớp với selected_ID
        try:
            if ma_sach != selected_ms:  # selected_ID phải được cập nhật trước khi sửa
                QMessageBox.critical(self.main_win_quanli, "Lỗi!", "Mã sách không thể chỉnh sửa.")
                self.uic.txt_ms.setText(selected_ms)  # Khôi phục lại mã sách gốc
                return

            # Thực thi câu lệnh SQL
            cursor.execute(update_sach, (ten_sach, int(so_luong), nxb, tac_gia, ma_sach))
            conn.commit()

            # Tải lại bảng sau khi sửa
            self.load_table_ql()
            QMessageBox.information(self.main_win_quanli, "Thành công", "Sửa thông tin sách thành công!")
        except Exception as e:
            # Xử lý lỗi nếu có
            QMessageBox.critical(self.main_win_quanli, "Lỗi!", f"Lỗi thực thi: {str(e)}")

    def xoasach(self):
        delete_sach = '''DELETE FROM Quanly WHERE Masach = ?'''
        cursor = conn.cursor()
        ma_sach = self.uic.txt_ms.text()

        if not ma_sach:  # Kiểm tra nếu `ma_sach` trống
            QMessageBox.critical(self.main_win_quanli, "Lỗi", "Vui lòng nhập mã sách muốn xóa!!")
            return

        try:
            # Thực hiện xóa sách
            cursor.execute(delete_sach, (ma_sach,))
            conn.commit()
            self.load_table_ql()  # Cập nhật lại bảng
            QMessageBox.information(self.main_win_quanli, "Thành công", "Xóa sách thành công!")
        except Exception as e:
            QMessageBox.critical(self.main_win_quanli, "Lỗi", f"Lỗi thực thi: {str(e)}")


    def timsach(self):
        timkiem = self.uic.txt_timkiemqly.text()
        if len(timkiem) == 0:
            self.load_table_ql()
        else:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM Quanly where Masach = ?",timkiem)
            result = cursor.fetchall()
            self.uic.tb_quanly.clearContents()  # clear trước khi cập nhật
            rowcount = len(result)
            self.uic.tb_quanly.setRowCount(rowcount)
            table_row = 0
            for row in result:
                self.uic.tb_quanly.setItem(table_row, 0, QTableWidgetItem(str(row[0])))
                self.uic.tb_quanly.setItem(table_row, 1, QTableWidgetItem(str(row[1])))
                self.uic.tb_quanly.setItem(table_row, 2, QTableWidgetItem(str(row[2])))
                self.uic.tb_quanly.setItem(table_row, 3, QTableWidgetItem(str(row[3])))
                self.uic.tb_quanly.setItem(table_row, 4, QTableWidgetItem(str(row[4])))
                table_row += 1
        self.uic.txt_timkiemqly.setText("")

class Thongke:
    def __init__(self):
        self.main_win_thongke = QMainWindow()
        self.uic = Ui_Thongke()
        self.uic.setupUi(self.main_win_thongke)
        self.uic.p_thongke.clicked.connect(self.thongke)
        self.uic.p_dx.clicked.connect(self.trangdn)
        self.uic.p_excel.clicked.connect(self.excel)
        self.uic.p_trolai.clicked.connect(self.qlai)

    def qlai(self):
        self.quanli = Quanly()
        self.quanli.show_ql()
        self.main_win_thongke.close()

    def trangdn(self):
        self.main_win = Dangnhap()
        self.main_win.show()
        self.main_win_thongke.close()

    def show_tk(self):
        self.main_win_thongke.show()

    def thongke(self):
        global results
        # Lấy ngày mượn và ngày trả từ giao diện
        ngay_muon = self.uic.Ngaymuon.date().toString("yyyy-MM-dd")  # Định dạng ngày tháng
        ngay_tra = self.uic.Ngaytra.date().toString("yyyy-MM-dd")  # Định dạng ngày tháng

        try:
            # Truy vấn cơ sở dữ liệu
            query = '''
                SELECT Tennguoimuon, Masach, Tensach, Soluong, NgayMuon, NgayTra, Lienhe
                FROM Muonvatra
                WHERE NgayMuon >= ? AND NgayTra <= ?
            '''
            cursor = conn.cursor()  # Tạo con trỏ cơ sở dữ liệu
            cursor.execute(query, (ngay_muon, ngay_tra))
            results = cursor.fetchall()  # Lấy tất cả kết quả truy vấn

            # Hiển thị kết quả lên bảng (QTableWidget)
            self.uic.tb_tk.setRowCount(len(results))  # Đặt số dòng
            self.uic.tb_tk.setColumnCount(7)  # Đặt số cột tương ứng với bảng Muonvatra
            self.uic.tb_tk.setHorizontalHeaderLabels([
                "Tên người mượn", "Mã sách", "Tên sách", "Số lượng", "Ngày mượn", "Ngày trả", "Liên hệ"
            ])  # Đặt tiêu đề cột

            # Điền dữ liệu vào bảng
            for row, record in enumerate(results):
                for col, item in enumerate(record):
                    self.uic.tb_tk.setItem(row, col, QTableWidgetItem(str(item)))

            # Thông báo nếu không có kết quả
            if not results:
                QMessageBox.information(self, "Kết quả", "Không có dữ liệu phù hợp với ngày mượn và ngày trả đã chọn.")
        except Exception as e:
            # Xử lý lỗi nếu có
            QMessageBox.critical(self, "Lỗi!", f"Lỗi thực thi: {str(e)}")

    def excel(self):
     try:
        ngay_muon = self.uic.Ngaymuon.date().toString("yyyy-MM-dd")  # Định dạng ngày tháng
        ngay_tra = self.uic.Ngaytra.date().toString("yyyy-MM-dd")  # Định dạng ngày tháng

         # Truy vấn cơ sở dữ liệu
        query = '''
                        SELECT Tennguoimuon, Masach, Tensach, Soluong, NgayMuon, NgayTra, Lienhe
                        FROM Muonvatra
                        WHERE NgayMuon >= ? AND NgayTra <= ?
                    '''
        cursor = conn.cursor()  # Tạo con trỏ cơ sở dữ liệu
        cursor.execute(query, (ngay_muon, ngay_tra))
        results = cursor.fetchall()  # Lấy tất cả kết quả truy vấn

        root = Tk()
        root.withdraw()  # Ẩn cửa sổ chính Tkinter
        file_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        if file_path:  # Nếu người dùng chọn một file
            # Tạo workbook và sheet mới
            wb = openpyxl.Workbook()
            ws = wb.active

        # Merge title và định dạng
        ws.merge_cells('A1:G1')
        ws['A1'] = "Thống kê ngày mượn và trả"
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        # Tiêu đề cột
        headers = ["Tên người mượn", "Mã sách", "Tên sách","Số lượng", "Ngày mượn", "Ngày trả", "Liên hệ"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ghi dữ liệu
        for i, row_data in enumerate(results, start=3):
            for j, value in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=value)

        # Thêm border cho tất cả các ô dữ liệu
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=len(results) + 2, min_col=1, max_col=7):
            for cell in row:
                cell.border = border

        # Lưu file Excel
        wb.save(file_path)

     except Exception as e:
         # Xử lý lỗi nếu có
         QMessageBox.critical(self, "Lỗi!", f"Lỗi thực thi: {str(e)}")
if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_win = Dangnhap()
    main_win.show()
    sys.exit(app.exec())